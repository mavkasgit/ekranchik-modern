"""
E2E тест: Матчинг OPC UA unload-событий с данными Excel.

Тестирует полный цикл:
1. Создаём тестовый Excel файл (лист "Подвесы")
2. Настраиваем excel_service на чтение тестового файла
3. Симулируем unload через opcua_service._cache → line_monitor._check_unload()
4. Вызываем GET /api/dashboard/opcua-unload-matched
5. Проверяем результат матчинга
"""
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch, AsyncMock

import pytest
import openpyxl
from httpx import AsyncClient, ASGITransport


# ---------------------------------------------------------------------------
# Вспомогательная функция для создания тестового Excel
# ---------------------------------------------------------------------------

def create_test_excel(filepath: str, rows: list[dict]) -> None:
    """
    Создаёт .xlsx файл с листом «Подвесы» и заполненными данными.

    Первые 2 строки — заголовки (skiprows=[0,1] при чтении pandas).
    Строка 3 — заголовок таблицы (будет прочитана как header).
    Данные начинаются со строки 4 (openpyxl row=4).

    Столбцы (1-indexed):
        D(4)=date, E(5)=number, F(6)=time,
        H(8)=material_type, I(9)=defect,
        K(11)=kpz_number, L(12)=client,
        M(13)=profile, Q(17)=color, T(20)=lamels_qty
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Подвесы"

    # Строки 1-2 — заголовки
    ws.cell(row=1, column=1, value="Заголовок")
    ws.cell(row=2, column=1, value="Подзаголовок")
    
    # Строка 3 — заголовок таблицы
    ws.cell(row=3, column=4, value="Дата")
    ws.cell(row=3, column=5, value="Номер")
    ws.cell(row=3, column=6, value="Время")

    for i, row_data in enumerate(rows, start=4):
        ws.cell(row=i, column=4, value=row_data.get("date"))               # D
        ws.cell(row=i, column=5, value=row_data.get("number"))             # E
        ws.cell(row=i, column=6, value=row_data.get("time"))               # F
        ws.cell(row=i, column=8, value=row_data.get("material_type", ""))  # H
        ws.cell(row=i, column=9, value=row_data.get("defect", ""))         # I
        ws.cell(row=i, column=11, value=row_data.get("kpz_number", ""))    # K
        ws.cell(row=i, column=12, value=row_data.get("client", ""))        # L
        ws.cell(row=i, column=13, value=row_data.get("profile", ""))       # M
        ws.cell(row=i, column=17, value=row_data.get("color", ""))         # Q
        ws.cell(row=i, column=20, value=row_data.get("lamels_qty", 0))     # T

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Хелперы
# ---------------------------------------------------------------------------

def _setup_excel(filepath: Path):
    """Настраивает excel_service на тестовый файл."""
    from app.services.excel_service import excel_service

    excel_service._active_file_name = str(filepath)
    excel_service._cache = None
    excel_service._cache_mtime = None
    excel_service._cache_path = None


def _setup_opcua_and_monitor():
    """Настраивает opcua_service и line_monitor для теста."""
    from app.services.opcua_service import opcua_service
    from app.services.line_monitor import line_monitor

    opcua_service._connected = True
    opcua_service._cache = {}

    line_monitor._bath34_pallete = None
    line_monitor._unload_events.clear()
    line_monitor._processed_unloads.clear()


def _save_state() -> dict:
    """Сохраняет текущее состояние сервисов для восстановления."""
    from app.services.excel_service import excel_service
    from app.services.opcua_service import opcua_service
    from app.services.line_monitor import line_monitor

    return {
        "active_file": excel_service._active_file_name,
        "cache": excel_service._cache,
        "cache_mtime": excel_service._cache_mtime,
        "cache_path": excel_service._cache_path,
        "connected": opcua_service._connected,
        "opcua_cache": opcua_service._cache.copy(),
        "bath34": line_monitor._bath34_pallete,
        "unload_events": list(line_monitor._unload_events),
        "processed": list(line_monitor._processed_unloads),
    }


def _restore_state(state: dict):
    """Восстанавливает состояние сервисов."""
    from app.services.excel_service import excel_service
    from app.services.opcua_service import opcua_service
    from app.services.line_monitor import line_monitor

    excel_service._active_file_name = state["active_file"]
    excel_service._cache = state["cache"]
    excel_service._cache_mtime = state["cache_mtime"]
    excel_service._cache_path = state["cache_path"]

    opcua_service._connected = state["connected"]
    opcua_service._cache = state["opcua_cache"]

    line_monitor._bath34_pallete = state["bath34"]
    line_monitor._unload_events.clear()
    for ev in state["unload_events"]:
        line_monitor._unload_events.append(ev)
    line_monitor._processed_unloads.clear()
    for pu in state["processed"]:
        line_monitor._processed_unloads.append(pu)


async def trigger_unload(hanger_id: int) -> None:
    """
    Имитирует выгрузку подвеса:
      1) Bath[34].Pallete = 0  → _check_unload() инициализирует _bath34_pallete=0
      2) Bath[34].Pallete = N  → _check_unload() детектирует переход 0→N
    Мокаем websocket_manager.broadcast чтобы не было ошибок.
    """
    from app.services.opcua_service import opcua_service
    from app.services.line_monitor import line_monitor

    with patch("app.services.line_monitor.websocket_manager") as mock_ws:
        mock_ws.broadcast = AsyncMock(return_value=0)
        mock_ws.connection_count = 0

        # Шаг 1: инициализация — палетта = 0
        opcua_service._cache["ns=4;s=Bath[34].Pallete"] = 0
        line_monitor._bath34_pallete = None
        await line_monitor._check_unload()  # bath34_pallete → 0

        # Шаг 2: подвес прибыл
        opcua_service._cache["ns=4;s=Bath[34].Pallete"] = hanger_id
        await line_monitor._check_unload()  # переход 0→N → _record_unload


async def call_matched_api(limit: int = 10) -> list[dict]:
    """GET /api/dashboard/opcua-unload-matched с мок-ом catalog_service."""
    from app.main import app

    with patch(
        "app.api.routes.dashboard.catalog_service"
    ) as mock_catalog:
        mock_catalog.get_profiles_photos_batch = AsyncMock(return_value={})

        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://test",
        ) as client:
            resp = await client.get(
                "/api/dashboard/opcua-unload-matched",
                params={"limit": limit},
            )
            assert resp.status_code == 200, (
                f"API вернул {resp.status_code}: {resp.text}"
            )
            return resp.json()


# ===========================================================================
# ТЕСТЫ
# ===========================================================================


async def test_correct_match_same_day(tmp_path: Path):
    """
    Подвес 42 с данными в Excel (дата=сегодня, 2 часа назад).
    После выгрузки API должен вернуть event с правильными полями.
    """
    now = datetime.now()
    entry_time = now - timedelta(hours=2)

    filepath = tmp_path / "test_same_day.xlsx"
    create_test_excel(str(filepath), [
        {
            "date": datetime(entry_time.year, entry_time.month, entry_time.day),
            "number": 42,
            "time": datetime(
                entry_time.year, entry_time.month, entry_time.day,
                entry_time.hour, entry_time.minute, 0,
            ),
            "material_type": "Пороги",
            "kpz_number": "101",
            "client": "ТестКТМ",
            "profile": "4040",
            "color": "серебро",
            "lamels_qty": 29,
        }
    ])

    state = _save_state()
    try:
        _setup_excel(filepath)
        _setup_opcua_and_monitor()

        await trigger_unload(42)
        data = await call_matched_api()

        assert len(data) == 1, f"Ожидался 1 event, получили {len(data)}"
        event = data[0]

        assert event["hanger"] == 42
        assert event["client"] == "ТестКТМ"
        assert event["profile"] == "4040"
        assert event["color"] == "серебро"
        assert event["lamels_qty"] == 29
        assert event["kpz_number"] == "101"
        assert event["material_type"] == "Пороги"

        # entry_time должен содержать корректное время (HH:MM)
        assert event["entry_time"] is not None
        assert ":" in event["entry_time"]
    finally:
        _restore_state(state)


async def test_correct_match_cross_day(tmp_path: Path):
    """
    Подвес 15 загружен вчера в 22:00.
    Выгрузка сегодня → API должен смотчить через переход дня.
    """
    yesterday = datetime.now() - timedelta(days=1)

    filepath = tmp_path / "test_cross_day.xlsx"
    create_test_excel(str(filepath), [
        {
            "date": datetime(yesterday.year, yesterday.month, yesterday.day),
            "number": 15,
            "time": datetime(yesterday.year, yesterday.month, yesterday.day, 22, 0, 0),
            "material_type": "Профиль",
            "kpz_number": "202",
            "client": "РП",
            "profile": "ЮП-1401",
            "color": "белый",
            "lamels_qty": 15,
        }
    ])

    state = _save_state()
    try:
        _setup_excel(filepath)
        _setup_opcua_and_monitor()

        await trigger_unload(15)
        data = await call_matched_api()

        assert len(data) == 1, f"Ожидался 1 event, получили {len(data)}"
        event = data[0]

        assert event["hanger"] == 15
        assert event["client"] == "РП"
        assert event["profile"] == "ЮП-1401"
        # entry_date должен быть вчерашним
        assert event["entry_date"] is not None
    finally:
        _restore_state(state)


async def test_no_match_different_hanger(tmp_path: Path):
    """
    Excel содержит только подвес 42.
    Триггерим выгрузку подвеса 99 — матч не найден.
    """
    now = datetime.now()
    entry_time = now - timedelta(hours=1)

    filepath = tmp_path / "test_no_match.xlsx"
    create_test_excel(str(filepath), [
        {
            "date": datetime(entry_time.year, entry_time.month, entry_time.day),
            "number": 42,
            "time": datetime(
                entry_time.year, entry_time.month, entry_time.day,
                entry_time.hour, entry_time.minute, 0,
            ),
            "material_type": "Пороги",
            "kpz_number": "101",
            "client": "ТестКТМ",
            "profile": "4040",
            "color": "серебро",
            "lamels_qty": 29,
        }
    ])

    state = _save_state()
    try:
        _setup_excel(filepath)
        _setup_opcua_and_monitor()

        await trigger_unload(99)
        data = await call_matched_api()

        assert len(data) == 1, f"Ожидался 1 event, получили {len(data)}"
        event = data[0]

        assert event["hanger"] == 99
        assert event["client"] == "—"
        assert event["profile"] == "—"
    finally:
        _restore_state(state)


async def test_closest_match_wins(tmp_path: Path):
    """
    Две записи для подвеса 42: 5 часов назад («Старый») и 1 час назад («Новый»).
    Матчинг должен выбрать ближайшую по времени → «Новый».
    """
    now = datetime.now()
    old_time = now - timedelta(hours=5)
    new_time = now - timedelta(hours=1)

    filepath = tmp_path / "test_closest.xlsx"
    create_test_excel(str(filepath), [
        {
            "date": datetime(old_time.year, old_time.month, old_time.day),
            "number": 42,
            "time": datetime(
                old_time.year, old_time.month, old_time.day,
                old_time.hour, old_time.minute, 0,
            ),
            "material_type": "Пороги",
            "kpz_number": "100",
            "client": "Старый",
            "profile": "1010",
            "color": "чёрный",
            "lamels_qty": 10,
        },
        {
            "date": datetime(new_time.year, new_time.month, new_time.day),
            "number": 42,
            "time": datetime(
                new_time.year, new_time.month, new_time.day,
                new_time.hour, new_time.minute, 0,
            ),
            "material_type": "Пороги",
            "kpz_number": "200",
            "client": "Новый",
            "profile": "2020",
            "color": "золото",
            "lamels_qty": 20,
        },
    ])

    state = _save_state()
    try:
        _setup_excel(filepath)
        _setup_opcua_and_monitor()

        await trigger_unload(42)
        data = await call_matched_api()

        assert len(data) == 1, f"Ожидался 1 event, получили {len(data)}"
        event = data[0]

        assert event["hanger"] == 42
        assert event["client"] == "Новый", (
            f"Ожидали клиента 'Новый' (ближайший), получили '{event['client']}'"
        )
        assert event["profile"] == "2020"
    finally:
        _restore_state(state)
