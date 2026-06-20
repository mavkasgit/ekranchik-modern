import asyncio
import logging
import os
import json
import time as time_mod
from datetime import datetime, time
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from asyncua import Server, ua

# Настройка логирования в консоль и в файл
backend_dir = Path(__file__).parent.resolve()
log_file = backend_dir / "custom_simulator.log"

# Настройка обработчиков логов
file_handler = logging.FileHandler(log_file, encoding='utf-8')
stream_handler = logging.StreamHandler()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[file_handler, stream_handler]
)
logger = logging.getLogger("CustomSimulator")
logging.getLogger("asyncua").setLevel(logging.WARNING)

logger.info(f"Логи симулятора записываются в {log_file}")

def get_excel_path() -> Path:
    """Определяет путь к активному файлу Excel на основе конфигурации бэкенда."""
    backend_dir = Path(__file__).parent.resolve()
    
    # Загружаем .env бэкенда
    env_path = backend_dir / ".env"
    if env_path.exists():
        load_dotenv(dotenv_path=env_path)
        logger.info(f"Загружен файл конфигурации: {env_path}")
    
    sim_enabled_str = os.getenv("SIMULATION_ENABLED", "true").lower()
    sim_enabled = sim_enabled_str in ("true", "1", "yes")
    
    # Ищем активный файл в static/active_excel_sim.json или active_excel.json
    static_dir = backend_dir / "static"
    config_filename = "active_excel_sim.json" if sim_enabled else "active_excel.json"
    config_path = static_dir / config_filename
    
    excel_file_path = None
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                excel_file_path = data.get("active_file")
                if excel_file_path:
                    logger.info(f"Найден активный файл из конфигурации: {excel_file_path}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла конфигурации {config_path}: {e}")
            
    if not excel_file_path:
        if sim_enabled:
            excel_file_path = os.getenv("EXCEL_TEST_FILE_PATH", "../testdata/Учет КПЗ 2026.xlsm")
        else:
            excel_file_path = os.getenv("EXCEL_REAL_FILE_PATH", "../testdata/Учет КПЗ 2026.xlsm")
        logger.info(f"Используется путь по умолчанию из .env: {excel_file_path}")

    excel_path = Path(excel_file_path)
    if not excel_path.is_absolute():
        excel_path = (backend_dir / excel_path).resolve()
        
    logger.info(f"Итоговый путь к Excel: {excel_path}")
    return excel_path

def load_real_templates(excel_path: Path) -> list[dict]:
    """
    Загружает реальные записи (не мусор от симулятора) из начала Excel файла
    для использования их в качестве шаблонов.
    """
    templates = []
    if not excel_path.exists():
        return templates
        
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        if "Подвесы" not in wb.sheetnames:
            wb.close()
            return templates
            
        ws = wb["Подвесы"]
        
        # Считываем первые 500 строк в поисках реальных данных
        for row in ws.iter_rows(min_row=4, max_row=500, max_col=20, values_only=True):
            if len(row) >= 20:
                val_d = row[3]  # D (Дата)
                val_e = row[4]  # E (Номер)
                client = row[11] # L (Клиент)
                profile = row[12] # M (Профиль)
                
                if val_d and val_e and client and profile:
                    client_str = str(client).lower()
                    # Игнорируем симуляторные записи
                    if "симулятор" not in client_str and "тест" not in client_str and client != "—":
                        templates.append({
                            "material_type": row[7],     # H
                            "kpz_number": row[10],       # K
                            "client": row[11],           # L
                            "profile": row[12],          # M
                            "color": row[16],            # Q
                            "lamels_qty": row[19]        # T
                        })
        wb.close()
    except Exception as e:
        logger.error(f"Ошибка загрузки шаблонов из Excel: {e}")
        
    if not templates:
        logger.warning("Не удалось загрузить шаблоны из файла. Используем встроенные реальные шаблоны на основе ТЗ.")
        templates = [
            {"material_type": "Пороги", "kpz_number": "КПЗ-146", "client": "Тест-КТМ", "profile": "ЮП-1401", "color": "белый", "lamels_qty": 16},
            {"material_type": "Фурнитура", "kpz_number": "КПЗ-147", "client": "АльфаПрофиль", "profile": "СП-777", "color": "золото", "lamels_qty": 17},
            {"material_type": "Профиль", "kpz_number": "КПЗ-148", "client": "Тест-КТМ", "profile": "Окно-15", "color": "черный", "lamels_qty": 18},
            {"material_type": "Пороги", "kpz_number": "КПЗ-149", "client": "АльфаПрофиль", "profile": "Порог-2", "color": "RAL 9005", "lamels_qty": 19},
            {"material_type": "Фурнитура", "kpz_number": "КПЗ-150", "client": "Тест-КТМ", "profile": "4040", "color": "серебро", "lamels_qty": 20},
        ]
        
    logger.info(f"Загружено {len(templates)} реальных шаблонов подвесов из Excel.")
    return templates

def get_last_hanger_and_next_row(excel_path: Path) -> tuple[int, int]:
    """
    Сканирует лист 'Подвесы' в Excel для нахождения последнего номера подвеса
    и первой пустой строки для записи с использованием быстрого последовательного чтения.
    """
    if not excel_path.exists():
        logger.warning(f"Файл Excel {excel_path} не найден. Используем ID подвеса по умолчанию: 800")
        return 799, 4
        
    try:
        # Открываем в read_only режиме для быстрого чтения большого файла
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        if "Подвесы" not in wb.sheetnames:
            logger.error(f"Лист 'Подвесы' не найден в {excel_path}. Используем дефолтные параметры.")
            wb.close()
            return 799, 4
            
        ws = wb["Подвесы"]
        
        last_hanger_id = 799
        row_index = 4
        
        # Сканируем строки последовательно, начиная со 4-й (min_row=4)
        for r_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=5, values_only=True), start=4):
            # В кортеже row: col D - индекс 3, col E - индекс 4
            if len(row) >= 5:
                val_d = row[3] # D
                val_e = row[4] # E
                if val_d is None and val_e is None:
                    row_index = r_idx
                    break
                if val_e is not None:
                    try:
                        last_hanger_id = int(val_e)
                    except (ValueError, TypeError):
                        pass
                row_index = r_idx + 1
            else:
                row_index = r_idx
                break
            
        wb.close()
        logger.info(f"Сканирование Excel завершено. Последний подвес: {last_hanger_id}, следующая строка для записи: {row_index}")
        return last_hanger_id, row_index
    except Exception as e:
        logger.error(f"Ошибка при сканировании Excel: {e}", exc_info=True)
        return 799, 4

def append_hanger_to_excel(excel_path: Path, hanger_id: int, row_index: int, now_dt: datetime, templates: list[dict], write_time: bool = False) -> tuple[int, int]:
    """
    Добавляет новую запись о подвесе в Excel в указанную строку, используя реальные шаблоны данных.
    По умолчанию записывает с пустым временем (статус: загрузка).
    Реализует retry-логику в случае блокировки файла (Errno 13).
    Возвращает (индекс_строки, индекс_следующей_строки).
    """
    if not excel_path.exists():
        logger.warning(f"Файл Excel не найден при записи: {excel_path}. Пропускаем запись.")
        return row_index, row_index + 1

    # Выбираем шаблон из реальных данных Excel
    template = templates[hanger_id % len(templates)]
    
    # Номер КПЗ генерируем на основе шаблона (или инкрементируем)
    base_kpz = template.get("kpz_number", "777")
    if base_kpz and str(base_kpz).startswith("КПЗ-"):
        try:
            kpz_num = int(str(base_kpz).split("-")[1])
            kpz_number = f"КПЗ-{kpz_num + 1}"
        except:
            kpz_number = f"КПЗ-{hanger_id}"
    else:
        kpz_number = f"КПЗ-{hanger_id}"

    max_retries = 5
    file_accessible = False
    
    for attempt in range(max_retries):
        try:
            # Быстрая проверка блокировки файла без полной загрузки openpyxl
            with open(excel_path, 'r+b'):
                pass
            file_accessible = True
            break
        except PermissionError as pe:
            logger.warning(f"[Попытка {attempt + 1}/{max_retries}] Excel файл заблокирован ({pe}). Повтор через 2 сек...")
            if attempt < max_retries - 1:
                time_mod.sleep(2)
            else:
                logger.error(f"Не удалось записать в Excel после {max_retries} попыток из-за блокировки файла.")
                return row_index, row_index + 1
        except Exception as e:
            logger.error(f"Ошибка проверки доступа к файлу: {e}")
            break

    if not file_accessible:
        return row_index, row_index + 1

    try:
        # Открываем для записи с сохранением VBA макросов (теперь гарантированно не будет PermissionError)
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        ws = wb["Подвесы"]
        
        # Для безопасности перепроверим, что строка действительно пустая
        row = row_index
        while True:
            val_d = ws.cell(row=row, column=4).value
            val_e = ws.cell(row=row, column=5).value
            if val_d is None and val_e is None:
                break
            row += 1
            
        # Записываем в ячейки
        ws.cell(row=row, column=4, value=now_dt.date())               # D: Дата
        ws.cell(row=row, column=5, value=hanger_id)                   # E: Номер
        # Если write_time=True, сразу пишем время, иначе оставляем пустым для отображения в "Загрузке"
        ws.cell(row=row, column=6, value=now_dt.time().replace(microsecond=0) if write_time else None) # F: Время
        ws.cell(row=row, column=8, value=template.get("material_type", "Профиль")) # H: Вид материала
        ws.cell(row=row, column=11, value=kpz_number)                 # K: Номер КПЗ
        ws.cell(row=row, column=12, value=template.get("client", "Тест-КТМ"))     # L: Клиент
        ws.cell(row=row, column=13, value=template.get("profile", "ЮП-1401"))    # M: Профиль
        ws.cell(row=row, column=17, value=template.get("color", "белый"))         # Q: Цвет
        ws.cell(row=row, column=20, value=template.get("lamels_qty", 10))         # T: Кол-во ламелей
        
        wb.save(excel_path)
        wb.close()
        logger.info(f"Excel обновлен (запись создана): Строка {row} | Подвес {hanger_id} | Профиль {template.get('profile')} | Клиент {template.get('client')}")
        return row, row + 1
    except Exception as e:
        logger.error(f"Не удалось записать в Excel: {e}", exc_info=True)
        return row_index, row_index + 1

def write_time_to_excel(excel_path: Path, row: int, now_dt: datetime) -> bool:
    """
    Записывает время старта (F: Время) в существующую строку Excel.
    Реализует быструю предварительную проверку блокировки.
    """
    if not excel_path.exists():
        return False
        
    max_retries = 5
    file_accessible = False
    
    for attempt in range(max_retries):
        try:
            with open(excel_path, 'r+b'):
                pass
            file_accessible = True
            break
        except PermissionError:
            time_mod.sleep(2)
            
    if not file_accessible:
        logger.error(f"Не удалось записать время старта в Excel для строки {row} после ретраев.")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        ws = wb["Подвесы"]
        ws.cell(row=row, column=6, value=now_dt.time().replace(microsecond=0)) # F: Время
        wb.save(excel_path)
        wb.close()
        logger.info(f"Excel обновлен (время записано): Строка {row} | Время {now_dt.strftime('%H:%M:%S')}")
        return True
    except Exception as e:
        logger.error(f"Ошибка записи времени в Excel: {e}")
        return False

async def run_simulator():
    server = Server()
    server.set_endpoint("opc.tcp://127.0.0.1:4840/freeopcua/server/")
    await server.init()
    
    # Register namespaces to get to index 4
    await server.register_namespace("urn:dummy:namespace2")
    await server.register_namespace("urn:dummy:namespace3")
    idx = await server.register_namespace("urn:omron:plc:namespace")  # namespace 4
    
    objects = server.get_objects_node()
    
    # Create power supply variables
    power_node_id = ua.NodeId("S8VK_X", idx)
    power_obj = await objects.add_object(power_node_id, "S8VK_X")
    power_vars = {
        'Status': await power_obj.add_variable(ua.NodeId("S8VK_X.Status", idx), "Status", True, varianttype=ua.VariantType.Boolean),
        'Voltage': await power_obj.add_variable(ua.NodeId("S8VK_X.Voltage", idx), "Voltage", 24.0, varianttype=ua.VariantType.Float),
        'Current': await power_obj.add_variable(ua.NodeId("S8VK_X.Current", idx), "Current", 5.0, varianttype=ua.VariantType.Float),
    }
    
    # Create Bath variables
    bath_vars = {}
    for bath_num in range(1, 41):
        bath_node_id = ua.NodeId(f"Bath[{bath_num}]", idx)
        bath_obj = await objects.add_object(bath_node_id, f"Bath[{bath_num}]")
        bath_vars[bath_num] = {
            'InUse': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].InUse", idx), "InUse", False, varianttype=ua.VariantType.Boolean),
            'Free': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].Free", idx), "Free", True, varianttype=ua.VariantType.Boolean),
            'Pallete': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].Pallete", idx), "Pallete", 0, varianttype=ua.VariantType.UInt32),
            'InTime': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].InTime", idx), "InTime", 0, varianttype=ua.VariantType.UInt32),
            'OutTime': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].OutTime", idx), "OutTime", 0, varianttype=ua.VariantType.UInt32),
            'dTime': await bath_obj.add_variable(ua.NodeId(f"Bath[{bath_num}].dTime", idx), "dTime", 0, varianttype=ua.VariantType.UInt32),
        }
        
    for bath_num in range(1, 41):
        for var in bath_vars[bath_num].values():
            await var.set_writable()
    for var in power_vars.values():
        await var.set_writable()
        
    await server.start()
    logger.info("Custom OPC UA Server started!")
    
    # Настройки симуляции Excel
    excel_path = get_excel_path()
    templates = load_real_templates(excel_path)
    last_hanger_id, next_row = get_last_hanger_and_next_row(excel_path)
    current_hanger_id = last_hanger_id + 1
    
    # Последовательность движения подвеса
    sequence = [3, 5, 7, 10, 17, 18, 19, 20, 31, 34]
    current_seq_idx = 0
    seconds_in_current_bath = 0
    bath_duration = 2  # 2 секунды в каждой ванне для быстрой демонстрации
    
    try:
        while True:
            # 1. Очищаем все ванны перед обновлением активной
            for bath_num in range(1, 41):
                await bath_vars[bath_num]['InUse'].write_value(False)
                await bath_vars[bath_num]['Free'].write_value(True)
                await bath_vars[bath_num]['Pallete'].write_value(ua.Variant(0, ua.VariantType.UInt32))
                await bath_vars[bath_num]['InTime'].write_value(ua.Variant(0, ua.VariantType.UInt32))
                await bath_vars[bath_num]['OutTime'].write_value(ua.Variant(0, ua.VariantType.UInt32))
                await bath_vars[bath_num]['dTime'].write_value(ua.Variant(0, ua.VariantType.UInt32))
            
            # 2. Обрабатываем активный цикл подвеса
            if current_seq_idx < len(sequence):
                # Если цикл только начинается, запишем новый подвес в Excel
                if current_seq_idx == 0 and seconds_in_current_bath == 0:
                    now = datetime.now()
                    row_idx, next_row = append_hanger_to_excel(excel_path, current_hanger_id, next_row, now, templates, write_time=False)
                    logger.info(f"Подвес {current_hanger_id} записан в Excel без времени на строке {row_idx}. Ждем 10 сек...")
                    await asyncio.sleep(10)
                    now_start = datetime.now()
                    write_time_to_excel(excel_path, row_idx, now_start)
                    logger.info(f"Время старта записано для подвеса {current_hanger_id} на строке {row_idx}.")
                
                active_bath = sequence[current_seq_idx]
                logger.info(f"Подвес {current_hanger_id} в Ванне {active_bath} ({seconds_in_current_bath}/{bath_duration}s)")
                
                await bath_vars[active_bath]['InUse'].write_value(True)
                await bath_vars[active_bath]['Free'].write_value(False)
                await bath_vars[active_bath]['Pallete'].write_value(ua.Variant(current_hanger_id, ua.VariantType.UInt32))
                await bath_vars[active_bath]['InTime'].write_value(ua.Variant(seconds_in_current_bath, ua.VariantType.UInt32))
                await bath_vars[active_bath]['OutTime'].write_value(ua.Variant(bath_duration, ua.VariantType.UInt32))
                await bath_vars[active_bath]['dTime'].write_value(ua.Variant(bath_duration, ua.VariantType.UInt32))
                
                seconds_in_current_bath += 1
                if seconds_in_current_bath >= bath_duration:
                    seconds_in_current_bath = 0
                    current_seq_idx += 1
            else:
                logger.info(f"Подвес {current_hanger_id} завершил маршрут. Подготовка следующего подвеса...")
                # Пауза перед запуском следующего цикла
                await asyncio.sleep(5)
                current_hanger_id += 1
                current_seq_idx = 0
                seconds_in_current_bath = 0
                
            await asyncio.sleep(1)
            
    finally:
        await server.stop()
        logger.info("Custom OPC UA Server stopped.")

if __name__ == "__main__":
    try:
        asyncio.run(run_simulator())
    except KeyboardInterrupt:
        logger.info("Симулятор остановлен пользователем.")
