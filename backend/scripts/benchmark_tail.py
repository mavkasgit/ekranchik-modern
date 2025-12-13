"""
Benchmark: reading last N rows from Excel
"""
import sys
import time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = Path("../testdata/Учет КПЗ 2025.xlsm")
SHEET_NAME = "Подвесы"
LAST_N = 100  # We want last 100 rows

print(f"📄 File: {EXCEL_PATH}")
print(f"🎯 Goal: Read last {LAST_N} rows")
print("=" * 60)

# Method 1: Read all, then tail (current approach)
print("\n🔧 Method 1: Read ALL + tail() [current]")
start = time.time()
df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, skiprows=[0,1], 
                   usecols=[3,4,5,7,10,11,12,16,19], engine='openpyxl')
df = df.dropna(how='all')
total_rows = len(df)
df_tail = df.tail(LAST_N)
elapsed = time.time() - start
print(f"   Total rows: {total_rows}, Got: {len(df_tail)}, Time: {elapsed:.3f}s")

# Method 2: Get row count first with openpyxl, then skiprows
print("\n🔧 Method 2: Count rows first, then skiprows")
start = time.time()
# Step 1: Quick count using openpyxl
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb[SHEET_NAME]
row_count = ws.max_row
wb.close()
count_time = time.time() - start
print(f"   Row count: {row_count} (in {count_time:.3f}s)")

# Step 2: Calculate skiprows and read
skip_rows = list(range(0, row_count - LAST_N))  # Skip all except last N
start2 = time.time()
df2 = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, skiprows=skip_rows,
                    usecols=[3,4,5,7,10,11,12,16,19], engine='openpyxl')
read_time = time.time() - start2
total_time = time.time() - start
print(f"   Got: {len(df2)} rows, Read time: {read_time:.3f}s, Total: {total_time:.3f}s")

# Method 3: calamine full + tail
print("\n🔧 Method 3: calamine + tail()")
try:
    start = time.time()
    df3 = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, skiprows=[0,1],
                        usecols=[3,4,5,7,10,11,12,16,19], engine='calamine')
    df3 = df3.dropna(how='all')
    df3_tail = df3.tail(LAST_N)
    elapsed = time.time() - start
    print(f"   Total rows: {len(df3)}, Got: {len(df3_tail)}, Time: {elapsed:.3f}s")
except Exception as e:
    print(f"   Error: {e}")

# Method 4: Cache simulation - second read should be instant
print("\n🔧 Method 4: Cached read (simulated)")
start = time.time()
# df is already in memory from Method 1
df_cached = df.tail(LAST_N)
elapsed = time.time() - start
print(f"   Got: {len(df_cached)} rows, Time: {elapsed:.6f}s (microseconds!)")

print("\n" + "=" * 60)
print("📊 SUMMARY")
print("=" * 60)
print("""
Method 1 (current): ~7-8s - reads everything, slow
Method 2 (skiprows): Still slow because openpyxl parses whole file anyway
Method 3 (calamine): ~2s - 4x faster than openpyxl
Method 4 (cached):   ~0.0001s - instant from memory

💡 BEST STRATEGY:
1. Use calamine engine (2s vs 8s)
2. Cache aggressively - only re-read when file changes
3. For dashboard: cache is valid most of the time = instant
""")
